import os
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# PY VER-1.0.3 APR|08|2025
print("\nRunning MPN Library Creator (Program 2)...\n")

# Paths and setup
TARGET_FOLDER = r'D:\MPNlibrary'
MPN_LIBRARY_FILE = os.path.join(TARGET_FOLDER, 'MPNLibrary.xlsx')
DUPLICATE_FILE = os.path.join(TARGET_FOLDER, 'Duplicate_MPNLibrary.xlsx')
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}
os.makedirs(TARGET_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# File selection
root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])

if file_path and allowed_file(file_path):
    try:
        df_new = pd.read_excel(file_path, usecols=['MPN', 'TYPE', 'CID', 'PKG', 'SDJ'])
        df_new['UPD'] = datetime.now().strftime("%d-%b-%Y|%I:%M %p")

        if os.path.exists(MPN_LIBRARY_FILE):
            df_existing = pd.read_excel(MPN_LIBRARY_FILE)

            # Drop UPD if duplicates from previous merge exist
            df_existing = df_existing.drop(columns=[col for col in df_existing.columns if col.startswith('UPD_')], errors='ignore')
            df_new = df_new.drop(columns=[col for col in df_new.columns if col.startswith('UPD_')], errors='ignore')

            # Perfect match rows
            merged = df_existing.merge(df_new, on=['MPN', 'TYPE', 'CID', 'PKG', 'SDJ'], how='inner')
            if not merged.empty:
                print(f"Found {len(merged)} exact duplicate rows. Skipping import of those.")

            # Remove perfect matches from new data
            df_new_unique = pd.concat([df_new, merged, merged]).drop_duplicates(keep=False)

            # Conflict check: same MPN, other details changed
            df_conflict = df_existing[df_existing['MPN'].isin(df_new_unique['MPN'])]
            df_conflict = df_conflict.merge(df_new_unique, on='MPN', suffixes=('_exist', '_new'))

            # Drop any UPD_x or UPD_y before comparing
            df_conflict = df_conflict.drop(columns=[col for col in df_conflict.columns if col.startswith('UPD_')], errors='ignore')

            df_conflict = df_conflict[
                (df_conflict['TYPE_exist'] != df_conflict['TYPE_new']) |
                (df_conflict['CID_exist'] != df_conflict['CID_new']) |
                (df_conflict['PKG_exist'] != df_conflict['PKG_new']) |
                (df_conflict['SDJ_exist'] != df_conflict['SDJ_new'])
            ]

            # Extract conflicting new rows
            conflicting_mpns = df_conflict['MPN'].unique()
            df_conflicting_rows = df_new_unique[df_new_unique['MPN'].isin(conflicting_mpns)]

            # Save to duplicate file
            if not df_conflicting_rows.empty:
                print("Conflicting MPN data found. Logging to Duplicate_MPNLibrary.xlsx")
                df_conflicting_rows.to_excel(DUPLICATE_FILE, index=False)

            # Keep truly new rows only
            df_to_append = df_new_unique[~df_new_unique['MPN'].isin(conflicting_mpns)]

            # Append new data
            df_final = pd.concat([df_existing, df_to_append], ignore_index=True)

            # Sort: conflicting MPNs on top
            df_conflicting_rows['__conflict__'] = True
            df_final['__conflict__'] = df_final['MPN'].isin(conflicting_mpns)
            df_final.sort_values(by='__conflict__', ascending=False, inplace=True)
            df_final.drop(columns='__conflict__', inplace=True)

        else:
            df_final = df_new
            conflicting_mpns = []

        # Save result to MPNLibrary
        df_final.to_excel(MPN_LIBRARY_FILE, index=False)

        # Conditional formatting for conflicting MPNs
        wb = load_workbook(MPN_LIBRARY_FILE)
        ws = wb.active
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for row in range(2, ws.max_row + 1):
            mpn_value = ws[f"A{row}"].value
            if mpn_value in conflicting_mpns:
                ws[f"A{row}"].fill = yellow_fill

        wb.save(MPN_LIBRARY_FILE)
        print(f'✅ Updated MPNLibrary saved to: {MPN_LIBRARY_FILE}')

    except Exception as e:
        print(f"❌ Error: {e}")
else:
    print("⚠️ No valid file selected.")

input("\nPress Enter to return to the menu...")
