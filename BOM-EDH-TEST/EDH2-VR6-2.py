import os
import sys
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

# PY VER-1.0.2 APR|08|2025

print("\nRunning MPN Library Creator (Program 2)...\n")

# Define target folder
TARGET_FOLDER = r'D:\MPNlibrary'
MPN_LIBRARY_FILE = os.path.join(TARGET_FOLDER, 'MPNLibrary.xlsx')
DUPLICATE_FILE = os.path.join(TARGET_FOLDER, 'Duplicate_MPNLibrary.xlsx')
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}
os.makedirs(TARGET_FOLDER, exist_ok=True)

# Check file extension
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Open file dialog
root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])

if file_path:
    if allowed_file(file_path):
        try:
            # Read input Excel file
            df_new = pd.read_excel(file_path, usecols=['MPN', 'TYPE', 'CID', 'PKG', 'SDJ'])
            timestamp = datetime.now().strftime("%d-%b-%Y|%I:%M %p")
            df_new['UPD'] = timestamp

            # Load existing library if available
            if os.path.exists(MPN_LIBRARY_FILE):
                df_existing = pd.read_excel(MPN_LIBRARY_FILE)

                # Find duplicates by MPN, TYPE, CID, PKG, SDJ
                df_duplicates = df_existing.merge(df_new, on=['MPN', 'TYPE', 'CID', 'PKG', 'SDJ'], how='inner')
                if not df_duplicates.empty:
                    print("Duplicate data found. Saving to Duplicate_MPNLibrary.xlsx")
                    df_duplicates.to_excel(DUPLICATE_FILE, index=False)

                # Append and drop perfect duplicates
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                df_combined.drop_duplicates(inplace=True)

                # Sort so duplicate MPNs come to the top
                dup_mpn_mask = df_combined.duplicated(subset='MPN', keep=False)
                df_combined['__DUP__'] = dup_mpn_mask.astype(int)
                df_combined.sort_values(by='__DUP__', ascending=False, inplace=True)
                df_combined.drop(columns='__DUP__', inplace=True)
            else:
                df_combined = df_new

            # Save to Excel
            df_combined.to_excel(MPN_LIBRARY_FILE, index=False)

            # Apply conditional formatting
            wb = load_workbook(MPN_LIBRARY_FILE)
            ws = wb.active
            mpn_col = 'A'
            last_row = ws.max_row

            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            dup_formula = f'COUNTIF(${mpn_col}$2:${mpn_col}${last_row},{mpn_col}2)>1'
            rule = FormulaRule(formula=[dup_formula], fill=yellow_fill)
            ws.conditional_formatting.add(f'{mpn_col}2:{mpn_col}{last_row}', rule)

            wb.save(MPN_LIBRARY_FILE)

            print(f'File saved to {MPN_LIBRARY_FILE}')

        except Exception as e:
            print(f'Error processing file: {str(e)}')
    else:
        print('Invalid file format. Only .xls and .xlsx allowed.')
else:
    print('No file selected.')

input("\nPress Enter to return to the menu...")
