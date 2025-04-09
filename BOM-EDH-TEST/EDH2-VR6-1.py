import os
import sys
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

print("\nRunning MPN Library Creator (Program 2)...\n")

# Define target folder
TARGET_FOLDER = r'D:\MPNlibrary'
MPN_LIBRARY_FILE = os.path.join(TARGET_FOLDER, 'MPNLibrary.xlsx')
DUPLICATE_FILE = os.path.join(TARGET_FOLDER, 'Duplicate_MPNLibrary.xlsx')
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}

# Ensure target folder exists
os.makedirs(TARGET_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Open file dialog
root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])

if file_path:
    if allowed_file(file_path):
        try:
            df_new = pd.read_excel(file_path, usecols=['MPN', 'TYPE', 'CID', 'PKG', 'SDJ'])

            # 🚫 Block if 'Sample' exists in any MPN
            if df_new['MPN'].str.lower().str.contains('sample').any():
                print("❌ 'Sample' MPN found in the input file. Please remove it before proceeding.")
                sys.exit()

            # Load or combine
            if os.path.exists(MPN_LIBRARY_FILE):
                df_existing = pd.read_excel(MPN_LIBRARY_FILE)
                df_duplicates = df_existing.merge(df_new, on=['MPN', 'TYPE', 'CID', 'PKG', 'SDJ'], how='inner')

                if not df_duplicates.empty:
                    print("Duplicate data found. Saving to Duplicate_MPNLibrary.xlsx")
                    df_duplicates.to_excel(DUPLICATE_FILE, index=False)

                df_combined = pd.concat([df_existing, df_new]).drop_duplicates()
            else:
                df_combined = df_new

            df_combined.to_excel(MPN_LIBRARY_FILE, index=False)

            # 🎨 Apply conditional formatting to highlight duplicate MPNs
            wb = load_workbook(MPN_LIBRARY_FILE)
            ws = wb.active

            # Add conditional formatting to column A (MPN column)
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            formula = f'=COUNTIF(A:A, A1)>1'
            rule = FormulaRule(formula=[formula], fill=fill)
            ws.conditional_formatting.add("A1:A1048576", rule)

            wb.save(MPN_LIBRARY_FILE)

            print(f'✅ File saved and formatted at: {MPN_LIBRARY_FILE}')
        except Exception as e:
            print(f'Error processing file: {str(e)}')
    else:
        print('Invalid file format. Only .xls and .xlsx allowed.')
else:
    print('No file selected.')

input("\nPress Enter to return to the menu...")

    
#sys.exit()
